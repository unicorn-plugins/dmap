"""
xlsx-build-sample.py
--------------------
DMAP 표준 XLSX 빌드 샘플 (xlsx-build-guide.md 4절 검증 규칙 9항 준수)

실행:    python xlsx-build-sample.py
산출물:  ./sample.xlsx (2시트 — 요약·상세 분기별 매출/비용/이익 + 합계 수식)
사전:    pip install openpyxl
"""

import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────
# 가이드 4-3. 스타일 객체 모듈 상단 1회 정의·재사용
# ─────────────────────────────────────────────────────────
KOR_FONT_NAME = "맑은 고딕"

TITLE_FONT = Font(name=KOR_FONT_NAME, size=14, bold=True, color="1F4E79")
HEADER_FONT = Font(name=KOR_FONT_NAME, size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name=KOR_FONT_NAME, size=10)
NUMBER_FONT = Font(name="Consolas", size=10)
TOTAL_FONT = Font(name=KOR_FONT_NAME, size=11, bold=True, color="1F4E79")

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

THIN = Side(border_style="thin", color="BFBFBF")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row_num, last_col):
    """헤더 행 스타일 일괄 적용."""
    ws.row_dimensions[row_num].height = 30
    for col in range(1, last_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER_ALL


def style_data_row(ws, row_num, last_col, *, alternate=False):
    """본문 행 스타일."""
    for col in range(1, last_col + 1):
        cell = ws.cell(row=row_num, column=col)
        if cell.font.name in (None, "Calibri"):
            cell.font = NUMBER_FONT if isinstance(cell.value, (int, float)) else BODY_FONT
        cell.alignment = RIGHT if isinstance(cell.value, (int, float)) else (CENTER if col == 1 else LEFT)
        cell.border = BORDER_ALL
        if alternate:
            cell.fill = ALT_FILL


def style_total_row(ws, row_num, last_col):
    ws.row_dimensions[row_num].height = 28
    for col in range(1, last_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.alignment = RIGHT if col > 1 else CENTER
        cell.border = BORDER_ALL


# ─────────────────────────────────────────────────────────
# 시트 1: 요약 (병합 셀 + 합계 수식)
# ─────────────────────────────────────────────────────────
def build_summary_sheet(wb):
    ws = wb.active
    ws.title = "요약"

    # 가이드 4-2. 병합 셀 1회 호출 (좌상단 셀에만 값)
    ws.merge_cells("A1:E1")
    ws["A1"] = "분기별 매출·비용·이익 요약"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 36

    headers = ["분기", "매출", "비용", "이익", "이익률(%)"]
    for col, value in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=value)
    style_header_row(ws, row_num=3, last_col=len(headers))

    data = [
        ("1Q", 100, 60),
        ("2Q", 120, 70),
        ("3Q", 150, 80),
        ("4Q", 180, 95),
    ]
    for i, (q, sales, cost) in enumerate(data):
        row_num = 4 + i
        ws.cell(row=row_num, column=1, value=q)
        ws.cell(row=row_num, column=2, value=sales)
        ws.cell(row=row_num, column=3, value=cost)
        # 가이드 4-4. 수식은 = 로 시작, 영문 대문자 셀 참조
        ws.cell(row=row_num, column=4, value=f"=B{row_num}-C{row_num}")
        ws.cell(row=row_num, column=5, value=f"=ROUND(D{row_num}/B{row_num}*100, 1)")
        style_data_row(ws, row_num, last_col=len(headers), alternate=(i % 2 == 1))

    # 합계 행
    total_row = 4 + len(data)
    ws.cell(row=total_row, column=1, value="합계")
    ws.cell(row=total_row, column=2, value=f"=SUM(B4:B{total_row - 1})")
    ws.cell(row=total_row, column=3, value=f"=SUM(C4:C{total_row - 1})")
    ws.cell(row=total_row, column=4, value=f"=SUM(D4:D{total_row - 1})")
    ws.cell(row=total_row, column=5, value=f"=ROUND(D{total_row}/B{total_row}*100, 1)")
    style_total_row(ws, total_row, last_col=len(headers))

    # 가이드 4-7. 열 너비
    widths = [10, 14, 14, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 가이드 4-5. 인쇄 설정
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:E{total_row}"


# ─────────────────────────────────────────────────────────
# 시트 2: 상세
# ─────────────────────────────────────────────────────────
def build_detail_sheet(wb):
    ws = wb.create_sheet("상세")

    headers = ["월", "제품", "매출(만원)", "수량", "단가"]
    for col, value in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=value)
    style_header_row(ws, row_num=1, last_col=len(headers))

    data = [
        ("01월", "A상품", 30, 60, "=ROUND(C2*10000/D2, 0)"),
        ("01월", "B상품", 25, 50, "=ROUND(C3*10000/D3, 0)"),
        ("02월", "A상품", 40, 80, "=ROUND(C4*10000/D4, 0)"),
        ("02월", "B상품", 35, 70, "=ROUND(C5*10000/D5, 0)"),
        ("03월", "A상품", 55, 100, "=ROUND(C6*10000/D6, 0)"),
        ("03월", "B상품", 45, 90, "=ROUND(C7*10000/D7, 0)"),
    ]
    for i, row in enumerate(data):
        row_num = 2 + i
        for col, value in enumerate(row, start=1):
            ws.cell(row=row_num, column=col, value=value)
        style_data_row(ws, row_num, last_col=len(headers), alternate=(i % 2 == 1))

    widths = [10, 14, 14, 10, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ─────────────────────────────────────────────────────────
# 가이드 4-8. 빌드 스크립트 진입점
# ─────────────────────────────────────────────────────────
def build():
    wb = Workbook()
    build_summary_sheet(wb)
    build_detail_sheet(wb)

    output = Path(__file__).parent / "sample.xlsx"
    wb.save(str(output))

    # 가이드 4-9. 자가 검증 (최소 항목)
    if not output.exists() or output.stat().st_size == 0:
        raise RuntimeError(f"산출물 검증 실패: {output}")

    print(f"[OK] saved: {output} ({output.stat().st_size:,} bytes, sheets: {len(wb.sheetnames)})")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(build())
    except Exception as e:
        print(f"[ERROR] {e}", file=sys.stderr)
        sys.exit(1)
